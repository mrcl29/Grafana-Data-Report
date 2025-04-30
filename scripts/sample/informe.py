#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# --------------------------------------------------------------------------------
#
#   Script para crear un informe excel (además de recojer datos en bruto en
#   archivos .cvs y .json) a partir de los datos monitorizados en Grafana.
#   El informe se crea a partir de la configuración introducida en el
#   archivo config.py en el mismo directorio que este script.
#
#   Autor: Marc Llobera Villalonga
#
# --------------------------------------------------------------------------------

import os
import requests
import json
import csv
import math
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import messagebox
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.chart import Reference, BarChart, LineChart
from openpyxl.chart.text import RichText
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.legend import Legend
from openpyxl.drawing.image import Image
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
import config

# Importamos la configuracion de config.py
ACTIVAR_SELECCION_RANGO_DE_FECHAS = config.ACTIVAR_SELECCION_RANGO_DE_FECHAS
DEBUG_FINAL = config.DEBUG_FINAL
IMG = ""
if config.IMG not in ["", None]:
    IMG = Image(config.IMG)
DATA_DIR = config.DATA_DIR
INFORMES_DIR = config.INFORMES_DIR
GRAFANA_URL = config.GRAFANA_URL
API_KEY = config.API_KEY
TITULO = config.TITULO
UIDS = config.UIDS
DASHBOARDS = config.DASHBOARDS

DAYS = 0
TIEMPO_INICIAL = timedelta(days=config.DAYS)
TIME_FINISH = datetime.utcnow().isoformat() + "Z"
TIME_START = (datetime.utcnow() - TIEMPO_INICIAL).isoformat() + "Z"

DATA_JSON_NAME = "query_data_"
DATA_CSV_NAME = "output_data_"

DEBUG_0 = DEBUG_FINAL
DEBUG_1 = DEBUG_FINAL
DEBUG_2 = DEBUG_FINAL
DEBUG_3 = DEBUG_FINAL  # POP-UP RANGO FECHAS

ANCHO_CELDA = 0.48  # cm
ALTO_CELDA = 0.53  # cm
ANCHOPX_CELDA = 18  # px
ALTOPX_CELDA = 20  # px

################################################# POP-UP RANGO DE FECHAS #################################################


def validar_fecha(fecha):
    patron = r'^(0[1-9]|[12][0-9]|3[01])/(0[1-9]|1[0-2])/\d{4}$'
    if re.match(patron, fecha):
        try:
            datetime.strptime(fecha, "%d/%m/%Y")
            return True
        except ValueError:
            return False
        # endtry
    # endif
    return False
# endfunction


def fecha_a_iso(fecha_str):
    fecha = datetime.strptime(fecha_str, "%d/%m/%Y")
    fecha = fecha.replace(hour=12, minute=0, second=0, microsecond=0)
    return fecha.isoformat() + "Z"
# endfunction


class FechaPopup:
    def __init__(self, master):
        self.master = master
        self.master.title("Seleccionar Rango")
        self.master.geometry("300x250")
        self.fecha_seleccionada_inicio = None
        self.fecha_seleccionada_final = None

        tk.Label(master, text="Introduce la fecha de inicio (dd/mm/aaaa):").pack()
        self.entrada_fecha_inicio = tk.Entry(master)
        self.entrada_fecha_inicio.pack()
        tk.Label(master, text="Introduce la fecha de fin (dd/mm/aaaa):").pack()
        self.entrada_fecha_final = tk.Entry(master)
        self.entrada_fecha_final.pack()

        tk.Button(master, text="Confirmar Fecha",
                  command=self.confirmar_fecha_manual).pack()

        tk.Button(master, text="1 dia",
                  command=lambda: self.seleccionar_fecha_predeterminada(1)).pack()
        tk.Button(master, text="3 dias",
                  command=lambda: self.seleccionar_fecha_predeterminada(3)).pack()
        tk.Button(master, text="7 dias",
                  command=lambda: self.seleccionar_fecha_predeterminada(7)).pack()
    # endfunction

    def confirmar_fecha_manual(self):
        fecha_inicio_str = self.entrada_fecha_inicio.get()
        fecha_final_str = self.entrada_fecha_final.get()
        if validar_fecha(fecha_inicio_str) and validar_fecha(fecha_final_str):
            fecha_inicio = datetime.strptime(fecha_inicio_str, "%d/%m/%Y")
            fecha_final = datetime.strptime(fecha_final_str, "%d/%m/%Y")
            if fecha_final <= fecha_inicio:
                messagebox.showerror(
                    "Error", "La fecha final debe ser posterior a la fecha de inicio.")
                return
            # endif
            self.fecha_seleccionada_inicio = fecha_a_iso(fecha_inicio_str)
            self.fecha_seleccionada_final = fecha_a_iso(fecha_final_str)
            self.master.quit()
        else:
            messagebox.showerror(
                "Error", "Formato de fecha incorrecto. Usa dd/mm/aaaa.")
        # endif
    # endfunction

    def seleccionar_fecha_predeterminada(self, dias):
        global DAYS
        DAYS = dias
        fecha_final = datetime.utcnow().isoformat() + "Z"
        fecha_inicio = (datetime.utcnow() -
                        timedelta(days=dias)).isoformat() + "Z"
        self.fecha_seleccionada_final = fecha_final
        self.fecha_seleccionada_inicio = fecha_inicio
        self.master.quit()
    # endfunction
# endclass


def mostrar_popup():
    global TIME_FINISH, TIME_START
    root = tk.Tk()
    popup = FechaPopup(root)
    root.mainloop()
    root.destroy()

    if popup.fecha_seleccionada_inicio and popup.fecha_seleccionada_final:
        TIME_START = popup.fecha_seleccionada_inicio
        TIME_FINISH = popup.fecha_seleccionada_final
        if DEBUG_3:
            print(
                f"La fecha seleccionada inicio es: {popup.fecha_seleccionada_inicio}")
            print(
                f"La fecha seleccionada final es: {popup.fecha_seleccionada_final}")
        # endif
    else:
        print("No se selecciono ninguna fecha.")
    # endif
# endfunction
##################################################################################################

################################################# Obtener datos de Grafana #################################################


def obtenerDatosGrafana(data_dir, grafana_url, api_key, dashboard_uid, panels):
    for panel_id in panels:
        url = f"{grafana_url}/api/dashboards/uid/{dashboard_uid}"
        header = {
            "Authorization": f"Bearer {api_key}",
            "Accept": "application/json",
            "Content-Type": "application/json",
        }
        query_data = None
        # Hacer la solicitud para obtener la metadata del dashboard
        response = requests.get(url, headers=header)

        # Verificar la respuesta
        if response.status_code == 200:
            data = response.json()

            # Buscar el panel con el PANEL_ID indicado
            dashboard_panels = data.get("dashboard", {}).get("panels", [])
            panel_metadata = next(
                (panel for panel in dashboard_panels if panel.get("id") == panel_id), None,)

            if panel_metadata:
                if DEBUG_0:
                    print(
                        f"\nMetadata del panel {panel_id}:", panel_metadata)
                # endif

                # Extraer la fuente de datos y las consultas del panel
                datasource = panel_metadata.get("datasource", None)
                targets = panel_metadata.get("targets", None)

                if datasource and targets:
                    # Ahora ejecutar la consulta contra el datasource usando la API de Grafana
                    datasource_URL = f"{grafana_url}/api/ds/query"

                    # Construir el payload para la consulta
                    query_payload = {
                        "queries": targets
                    }

                    for query in query_payload["queries"]:
                        if "query" in query:
                            # Reemplazar el rango de tiempo en la query
                            query["query"] = str(query["query"]).replace(
                                "range(start: v.timeRangeStart, stop:v.timeRangeStop)", f"range(start: {TIME_START}, stop:{TIME_FINISH})"
                            )
                            if DAYS != 0:
                                frc = math.ceil((DAYS*24*60)/1001)
                                # No son datos binarios
                                if (not panels.get(panel_id)[3]):
                                    query["query"] = str(query["query"]).replace(
                                        "aggregateWindow(every: 10s, fn: last)", f"aggregateWindow(every: {frc}m, fn: last)"
                                    )
                                # endif
                            # endif
                        # endif
                    # endfor

                    if DEBUG_0:
                        print(f"\Query: {str(query)}")
                    # endif

                    # Ejecutar la consulta
                    query_response = requests.post(
                        datasource_URL, headers=header, json=query_payload)

                    # Verificar la respuesta de la consulta
                    if query_response.status_code == 200 or query_response.status_code == 400:
                        query_data = query_response.json()

                        # Guardar el JSON en un archivo
                        file_path_data_json = os.path.join(
                            data_dir, DATA_JSON_NAME + panels.get(panel_id)[0]+".json")
                        with open(file_path_data_json, "w") as archivo:
                            json.dump(query_data, archivo)
                        if DEBUG_1:
                            print("\nRespuesta obtenida correctamente")
                            # El json obtenido
                            # Convertir el JSON a una cadena formateada
                            json_str = json.dumps(query_data, indent=2)

                            # Dividir la cadena en lineas
                            lines = json_str.splitlines()

                            # Tomar las primeras 50 lineas
                            first_50_lines = lines[:50]

                            # Unir las lineas de nuevo en una cadena
                            result = "\n".join(first_50_lines)
                            print("\nDatos obtenidos de la consulta:", result)
                        # endif
                    else:
                        print(
                            f"\nError {query_response.status_code}: {query_response.text}")
                    # endif
                else:
                    print(
                        "\nNo se encontro el datasource o las consultas asociadas al panel.")
                # endif
            else:
                print(f"\nNo se encontro el panel con ID {panel_id}")
            # endif
        else:
            print(
                f"\nError {response.status_code}: No se pudieron obtener los datos del dashboard")
        # endif
# endfunction
##################################################################################################

################################################# Excel de Datos #################################################


def convert_timestamp(timestamp):
    # Convertir de milisegundos a segundos
    timestamp_seconds = int(timestamp) / 1000
    # Convertir a objeto datetime
    dt_object = datetime.utcfromtimestamp(timestamp_seconds)
    # Formatear como string
    return dt_object.strftime("%Y-%m-%d %H:%M:%S")
# endfunction


def excelDeDatos(data_dir, panels):
    for panel_id in panels.keys():
        file_path_data_json = os.path.join(
            data_dir, DATA_JSON_NAME + panels.get(panel_id)[0]+".json")

        # Abrir y leer el archivo JSON
        with open(file_path_data_json, "r") as archivo:
            query_data = json.load(archivo)
        # endwith

        # Definir el archivo CSV donde guardar todos los datos
        file_path_data_csv = os.path.join(
            data_dir, DATA_CSV_NAME + panels.get(panel_id)[0] + ".csv")

        # Diccionario para almacenar los datos de todas las mediciones
        all_data = {}
        measurements = set()

        # Procesar cada conjunto de datos: "A", "B", "C", etc.
        for query_key in query_data["results"].keys():
            frames = query_data["results"][query_key].get("frames", [])

            if frames:
                for frame in frames:
                    values = frame["data"]["values"]
                    measurement_name = frame["schema"]["meta"]["executedQueryString"].split(
                        'r._measurement == ')[1].split(')')[0].replace('"', '')
                    measurements.add(measurement_name)

                    for time, value in zip(values[0], values[1]):
                        if value is not None:
                            converted_time = convert_timestamp(time)
                            aux = datetime.strptime(
                                converted_time, "%Y-%m-%d %H:%M:%S")
                            converted_time_aux = aux.strftime("%d-%m-%Y %H:%M")

                            if converted_time_aux not in all_data:
                                all_data[converted_time_aux] = {}
                            # endif
                            all_data[converted_time_aux][measurement_name] = int(
                                value)
                        # endif
                    # endfor
                # endfor

                if DEBUG_2:
                    print(f"\nDatos procesados para el conjunto {query_key}.")
                # endif
            else:
                print(f"\nNo se encontraron datos en el conjunto {query_key}.")
            # endif
        # endfor

        # Escribir los datos en el CSV
        with open(file_path_data_csv, mode="w", newline="", encoding="utf-8") as file:
            writer = csv.writer(file, delimiter=";", quoting=csv.QUOTE_MINIMAL)

            # Escribir encabezado
            header = ["Time"] + list(measurements)
            writer.writerow(header)

            # Escribir datos
            for time in sorted(all_data.keys()):
                row = [time]
                for measurement in measurements:
                    row.append(all_data[time].get(measurement, ""))
                # endfor
                writer.writerow(row)
            # endfor
        # endwith

        if DEBUG_2:
            print(
                f"\nDatos exportados al CSV para el conjunto {query_key}.")
        # endif
    # endfor
# endfunction
##################################################################################################


################################################# Crear Informe #################################################
# CONSTANTES
columna_tiempo = 1
row_start = 2  # Filas de datos inician en la segunda fila
col_start = 2

# Funcion para insertar un encabezado


def insert_header(worksheet, text, row, start_col, end_col):
    # Combinar celdas para el encabezado
    worksheet.merge_cells(
        start_row=row, start_column=start_col, end_row=row, end_column=end_col)

    # Obtener la celda combinada
    cell = worksheet.cell(row=row, column=start_col)

    # Establecer el valor y el estilo
    cell.value = text
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
# endfunction


def nuevaHoja(wb, data_path, name, evitarDatosVacios=True):
    ### Introducir Datos ###
    # Crear una nueva hoja para los datos filtrados
    ws_filtered = wb.create_sheet(name)
    with open(data_path, mode="r", newline="", encoding="utf-8") as archivo_csv:
        # Crear un objeto lector CSV
        lector_csv = csv.reader(
            archivo_csv, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        # Copiar el encabezado
        header = next(lector_csv)  # Lee la primera linea como encabezado
        ws_filtered.append(header)
        # Filtrar y copiar los datos
        rows_to_include = set()
        # Start with row index 2 for filtering
        for i, row in enumerate(lector_csv):
            if (evitarDatosVacios):
                add = True
                for j in range(1, len(row)):
                    if row[j] == '0' or row[j] == '' or row[j] == None or row[j] == 0 or row[j] == ' ':
                        add = False
                    # endif
                # endfor
                if (add):
                    rows_to_include.add(i - 1)  # Fila anterior
                    rows_to_include.add(i)       # Fila actual
                    rows_to_include.add(i + 1)   # Fila siguiente
                # endif
            else:
                rows_to_include.add(i)       # Fila actual
            # endif
        # endfor

        # Volver a leer el archivo para copiar las filas seleccionadas a la nueva hoja
        archivo_csv.seek(0)  # Volver al inicio del archivo
        next(lector_csv)  # Saltar el encabezado
        for i, row in enumerate(lector_csv):
            if i in rows_to_include:
                # Convertir a float
                for j in range(1, len(row)):
                    try:
                        if row[j] == '' or row[j] == '' or row[j] == None or row[j] == ' ':
                            row[j] = float(0)
                            break
                        # endif
                        row[j] = float(int(row[j]))
                    except ValueError:
                        row[j] = 0  # Or any default value you prefer
                    # endtry
                # endfor
                ws_filtered.append(row)
            # endif
        # endfor
    # endwith
    return ws_filtered
    ######
# endfunction


# CONSTANTES PARA LOS GRAFICOS
anchura_linea = 9000


def crear_grafico(chart, hoja, name, pos_x, pos_y, ancho, alto, binario=False, leyenda=False, series_colors=["ff0000", "000000", "a95700", "27fc00", "003dfc", "df00fc"]):
    ###### Crear el grafico ######
    # Anadir series de datos al grafico
    color_i = -1
    max_row = hoja.max_row
    max_col = hoja.max_column

    for j in range(2, max_col+1):
        # Obtener el nombre del conjunto (measurement)
        measurement = hoja.cell(row=1, column=j).value

        # Rango de datos de este conjunto
        data_range = Reference(hoja, min_col=j,
                               min_row=row_start, max_row=max_row)
        categories = Reference(hoja, min_col=columna_tiempo,
                               min_row=row_start, max_row=max_row)

        # Anadir serie de datos al grafico
        chart.add_data(data_range, titles_from_data=False)
        if hoja.max_row > 1:
            chart.set_categories(categories)
        # endif

        color_i = color_i + 1
        if color_i == len(series_colors):
            color_i = 0
        # endif

        # Personalizar la serie
        s = chart.series[-1]  # ultima serie anadida
        s.graphicalProperties.line.width = anchura_linea  # Anchura de la linea
        s.graphicalProperties.solidFill = ColorChoice(
            srgbClr=series_colors[color_i])
        s.graphicalProperties.line.solidFill = ColorChoice(
            srgbClr=series_colors[color_i])  # Borde
        s.marker.graphicalProperties.solidFill = ColorChoice(
            srgbClr=series_colors[color_i])
        s.marker.graphicalProperties.line.solidFill = ColorChoice(
            srgbClr=series_colors[color_i])
        chart.gapWidth = 150  # Aumenta el espacio entre barras, haciendolas mas delgadas
        chart.overlap = -25

        # Etiqueta para la leyenda
        if leyenda:
            s.title = SeriesLabel(v=measurement)
        # endif
    # endif

    # Establecer y formatear el titulo
    chart.title = name
    chart.title.tx.rich.paragraphs[0].r[0].rPr = CharacterProperties(
        sz=1100, b=True, solidFill=ColorChoice(srgbClr="808080"))

    # Personalizar el eje X
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(
        defRPr=CharacterProperties(sz=500, solidFill=ColorChoice(srgbClr="808080"))))])
    chart.x_axis.scaling.gap = 50
    chart.x_axis.majorTickMark = "none"
    chart.x_axis.minorTickMark = "none"
    chart.x_axis.txPr.properties.rot = "0"

    # Personalizar el eje Y
    if binario:
        chart.y_axis.crosses = "min"
        chart.y_axis.min = 0.00
        chart.y_axis.max = 1.00
        chart.y_axis.minorUnit = 1.0
        chart.y_axis.majorUnit = 0.0
    # endif
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(
        defRPr=CharacterProperties(sz=700, solidFill=ColorChoice(srgbClr="808080"))))])
    chart.y_axis.majorTickMark = "none"  # Eliminar las marcas grandes en el eje Y
    chart.y_axis.minorTickMark = "none"  # Eliminar las marcas pequenas en el eje Y

    # Estilo y dimensiones del grafico
    chart.style = 1
    if leyenda:
        chart.legend = Legend()
        chart.legend.position = "t"  # Colocar la leyenda a la derecha
        chart.legend.overlay = False  # No superponer la leyenda con el grafico
        chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(
            defRPr=CharacterProperties(sz=800, solidFill=ColorChoice(srgbClr="000000"))))])
    else:
        chart.legend = None
    # endif

    # Maximizar el area de trazado y reducir el margen izquierdo
    chart.plot_area.layout = Layout(
        ManualLayout(
            x=0.02,
            y=0.1,
            h=0.85,
            w=0.98
        )
    )

    # Ajustar tamano y posicion
    pos = XDRPoint2D(x=pos_x, y=pos_y)
    width = pixels_to_EMU(ancho)
    height = pixels_to_EMU(alto)
    size = XDRPositiveSize2D(cx=width, cy=height)
    anchor = AbsoluteAnchor(pos=pos, ext=size)
    chart.anchor = anchor

    ###### Retornar el grafico ######
    return chart
# endfunction


# CONSTANTES PARA INFORME
col_width = 2.43
row_height = 14.58
global NUM_ROWS_SHEET
NUM_ROWS_SHEET = 48

# Definir el estilo del borde
border_style = Side(style='thin')
# Crear el objeto Border con el estilo en todos los lados
border = Border(left=border_style,
                right=border_style,
                top=border_style,
                bottom=border_style)
font_style = Font(size=8)  # Tamano de fuente 8
alignment_style = Alignment(horizontal='center', vertical='center')


def apply_style_to_range(sheet, cell_range):
    for row in sheet[cell_range]:
        for cell in row:
            cell.font = font_style
            cell.alignment = alignment_style
            cell.border = border
        # endfor
    # endfor
# endfunction


def createMAXMIN(ws_source, ws_destination, idx):
    i = idx
    for col_idx, col in enumerate(ws_source.iter_cols(min_col=2)):
        max_value = float('-inf')
        min_value = float('inf')
        max_time = None
        min_time = None

        for row_idx, row in enumerate(ws_source.iter_rows(min_row=2)):
            time = row[0].value
            value = col[row_idx].value
            if isinstance(value, (int, float)):
                if value > max_value:
                    max_value = value
                    max_time = time
                # endif
                if value < min_value:
                    min_value = value
                    min_time = time
                # endif
            # endif
        # endfor

        # Formatear los valores con dos decimales
        max_value_formatted = f"{max_value:.2f}" if max_value != float(
            '-inf') else "N/A"
        min_value_formatted = f"{min_value:.2f}" if min_value != float(
            'inf') else "N/A"

        ws_destination[f"A{str(i)}"] = "MAX_" + \
            str(col[0].value)
        ws_destination[f"A{str(i+1)}"] = str(max_value_formatted)
        ws_destination[f"A{str(i+2)}"] = str(max_time)
        ws_destination[f"B{str(i)}"] = "MIN_" + \
            str(col[0].value)
        ws_destination[f"B{str(i+1)}"] = str(min_value_formatted)
        ws_destination[f"B{str(i+2)}"] = str(min_time)

        i += 4
    # endfor
# endfunction


def informe(titulo, dashboards):
    # Crear un nuevo libro y hoja
    wb = Workbook()

    ws = wb.active
    # Cambiar la vista a "Diseno de pagina"
    ws.sheet_view.view = "pageLayout"
    # Ocultar las lineas de division
    ws.sheet_view.showGridLines = False
    # Configuracion de la disposicion de la pagina
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4  # Tamano de pagina A4
    ws.title = "Informe"
    # Anadir encabezado
    ws.oddHeader.center.text = titulo
    ws.oddHeader.center.size = 14
    ws.oddHeader.center.font = "Arial,Bold"

    insert_header(ws, "", 1, 1, 35)

    # Ajustar el tamano de las filas y columnas
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = col_width
    # endfor
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = row_height
    # endfor

    # Crear fechas de inicio y fin
    formatted_start = (datetime.fromisoformat(
        TIME_START.replace("Z", "+00:00"))).strftime("%d/%m/%Y %H:%M")
    time_finish_aux = (datetime.fromisoformat(
        TIME_FINISH.replace("Z", "+00:00")))
    formatted_finish = time_finish_aux.strftime("%d/%m/%Y %H:%M")
    ws["E2"] = "Fecha INICIO: "
    ws.merge_cells("K2:Q2")
    ws["K2"] = formatted_start
    ws["E3"] = "Fecha FIN: "
    ws.merge_cells("K3:Q3")
    ws["K3"] = formatted_finish

    # Ajustar el tamano de la imagen
    if IMG not in [None, ""]:
        IMG.width = 170
        IMG.height = 85
        ws.add_image(IMG, "Z1")

    EMU_ANCHO_TOTAL = ANCHOPX_CELDA*34
    EMU_ANCHO_TOTAL_PEQUEnO = (EMU_ANCHO_TOTAL/2)
    EMU_ALTURA_PEQUEnO = (ALTOPX_CELDA/2)*16
    EMU_ALTURA_MEDIANO = (ALTOPX_CELDA/4)*20
    EMU_ALTURA_GRANDE = EMU_ALTURA_PEQUEnO*2
    x_offset = pixels_to_EMU(ANCHOPX_CELDA/2)  # Offset horizontal en EMUs
    x_offset_mitad = x_offset*2*18
    y_offset = pixels_to_EMU(ALTOPX_CELDA/4)   # Offset vertical en EMUs
    casella_EMU = pixels_to_EMU(ALTOPX_CELDA)

    ws_datos = wb.create_sheet("DATOS")

    pos_actual = 4
    suma_pequeno = 9
    suma_mediano = 6
    suma_grande = 17

    def evitarCorte(l, b):
        global NUM_ROWS_SHEET
        if b:
            x = NUM_ROWS_SHEET - 1
        else:
            x = NUM_ROWS_SHEET
        resto = pos_actual % x
        if resto > x-l:
            return x-resto+1
        # endif
        if pos_actual > 95 - suma_grande and pos_actual < 130:
            return 1
        else:
            return 0
        # endif
    # endfunction

    b = False
    for nombre, paneles in dashboards.items():

        z = evitarCorte(2, b)
        if (z == 0):
            pos_actual += 1
        else:
            pos_actual += zip
        # endif

        ws[f"C{str(pos_actual)}"] = nombre[1]
        ws.merge_cells(f"C{str(pos_actual)}:AG{str(pos_actual)}")
        color_relleno = PatternFill(
            start_color="6BB9AE", end_color="6BB9AE", fill_type="solid")
        ws[f"C{str(pos_actual)}"].fill = color_relleno
        ws[f"C{str(pos_actual)}"].font = Font(color="FFFFFF", bold=True)
        pos_actual += 1

        pan_p = 0
        series_colors = ["FF0000", "00FF00", "0000FF",
                         "00FFFF", "FF00FF", "FFA500", "800080"]
        series_colors_aux = ["ff4d4d", "4dff4d", "4d4dff",
                             "4dffff", "ff4dff", "ffc04d", "cd00cd"]
        # CLAVE:[NOMBRE, TIPO(LINEAS, BARRAS), TAMAnO(PEQUEnO, MEDIANO, GRANDE), BINARIO(True, False), LEYENDA(True, False), EXTRA(MAXMIN, INFO), EXTRA_INFO(mensaje)]
        for id, valores in paneles.items():
            if (pos_actual > NUM_ROWS_SHEET):
                b = True
            # endif

            file_path_data_csv = os.path.join(
                DATA_DIR+"/"+nombre[0]+"/", DATA_CSV_NAME+valores[0]+".csv")
            ws_data = nuevaHoja(
                wb, file_path_data_csv, f"Raw Data {nombre[0]}")

            if valores[1] == "L":
                tipo = LineChart()
            elif valores[1] == "B":
                tipo = BarChart()
            # endif

            x = x_offset
            pos_derecha = False
            if valores[2] == "P":
                t = 0
                if valores[5] == "INFO":
                    t = 1
                elif valores[5] == "MAXMIN":
                    t = 2
                # endif
                pos_actual += evitarCorte(suma_pequeno+t, b)
                ancho = EMU_ANCHO_TOTAL_PEQUEnO
                alto = EMU_ALTURA_PEQUEnO
                y = y_offset+casella_EMU*pos_actual
                if pan_p == 0:
                    pan_p += 1
                elif pan_p == 1:
                    x = x_offset_mitad
                    pan_p = 0
                    pos_derecha = True
                    pos_actual += suma_pequeno
                # endif
            elif valores[2] == "M":
                t = 0
                if valores[5] == "INFO":
                    t = 1
                elif valores[5] == "MAXMIN":
                    t = 2
                # endif
                pos_actual += evitarCorte(suma_mediano+t, b)
                ancho = EMU_ANCHO_TOTAL
                alto = EMU_ALTURA_MEDIANO
                if pan_p == 1:
                    pan_p = 0
                    pos_actual += suma_pequeno
                # endif
                y = y_offset+casella_EMU*pos_actual
                pos_actual += suma_mediano
            elif valores[2] == "G":
                t = 0
                if valores[5] == "INFO":
                    t = 1
                elif valores[5] == "MAXMIN":
                    t = 2
                # endif
                pos_actual += evitarCorte(suma_grande+t, b)
                ancho = EMU_ANCHO_TOTAL
                alto = EMU_ALTURA_GRANDE
                if pan_p == 1:
                    pan_p = 0
                    pos_actual += suma_pequeno
                # endif
                y = y_offset+casella_EMU*pos_actual
                pos_actual += suma_grande
            # endif

            chart = crear_grafico(
                tipo, ws_data, valores[0], x, y, ancho, alto, binario=valores[3], leyenda=valores[4], series_colors=series_colors)

            # Anadir el grafico a la hoja de trabajo
            ws.add_chart(chart)

            indice_datos = 1

            if valores[5] == "INFO":
                pos_actual += 1
                if ws_data.max_row < 2:
                    ws.merge_cells(f"C{str(pos_actual)}:AG{str(pos_actual)}")
                    ws[f"C{str(pos_actual)}"] = valores[6]
                    ws[f"C{str(pos_actual)}"].alignment = Alignment(
                        horizontal='center', vertical='center')
                # endif
            # endif
            elif valores[5] == "MAXMIN":
                createMAXMIN(ws_source=ws_data,
                             ws_destination=ws_datos, idx=indice_datos)

                if not pos_derecha:
                    pos_actual_aux = pos_actual + suma_pequeno + 1
                    ws.merge_cells(
                        f"A{str(pos_actual_aux)}:G{str(pos_actual_aux)}")
                    ws[f"A{str(pos_actual_aux)}"] = valores[6] + " max.: "
                    ws[f"A{str(pos_actual_aux)}"].alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.merge_cells(
                        f"H{str(pos_actual_aux)}:I{str(pos_actual_aux)}")
                    ws[f"H{str(pos_actual_aux)}"] = ws_datos[f"A{str(indice_datos+1)}"].value
                    ws[f"J{str(pos_actual_aux)}"] = valores[7]
                    ws.merge_cells(
                        f"K{str(pos_actual_aux)}:P{str(pos_actual_aux)}")
                    ws[f"K{str(pos_actual_aux)}"] = ws_datos[f"A{str(indice_datos+2)}"].value
                    ws[f"K{str(pos_actual_aux)}"].alignment = Alignment(
                        horizontal='center', vertical='center')

                    pos_actual_aux += 1

                    ws.merge_cells(
                        f"A{str(pos_actual_aux)}:G{str(pos_actual_aux)}")
                    ws[f"A{str(pos_actual_aux)}"] = valores[6] + " min.: "
                    ws[f"A{str(pos_actual_aux)}"].alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.merge_cells(
                        f"H{str(pos_actual_aux)}:I{str(pos_actual_aux)}")
                    ws[f"H{str(pos_actual_aux)}"] = ws_datos[f"B{str(indice_datos+1)}"].value
                    ws[f"J{str(pos_actual_aux)}"] = valores[7]
                    ws.merge_cells(
                        f"K{str(pos_actual_aux)}:P{str(pos_actual_aux)}")
                    ws[f"K{str(pos_actual_aux)}"] = ws_datos[f"B{str(indice_datos+2)}"].value
                    ws[f"K{str(pos_actual_aux)}"].alignment = Alignment(
                        horizontal='center', vertical='center')
                else:
                    pos_derecha = False
                    pos_actual += 1

                    ws.merge_cells(
                        f"S{str(pos_actual)}:Y{str(pos_actual)}")
                    ws[f"S{str(pos_actual)}"] = valores[6] + " max.: "
                    ws[f"S{str(pos_actual)}"].alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.merge_cells(
                        f"Z{str(pos_actual)}:AA{str(pos_actual)}")
                    ws[f"Z{str(pos_actual)}"] = ws_datos[f"A{str(indice_datos+1)}"].value
                    ws[f"AB{str(pos_actual)}"] = valores[7]
                    ws.merge_cells(
                        f"AC{str(pos_actual)}:AH{str(pos_actual)}")
                    ws[f"AC{str(pos_actual)}"] = ws_datos[f"A{str(indice_datos+2)}"].value
                    ws[f"AC{str(pos_actual)}"].alignment = Alignment(
                        horizontal='center', vertical='center')

                    pos_actual += 1

                    ws.merge_cells(
                        f"S{str(pos_actual)}:Y{str(pos_actual)}")
                    ws[f"S{str(pos_actual)}"] = valores[6] + " min.: "
                    ws[f"S{str(pos_actual)}"].alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.merge_cells(
                        f"Z{str(pos_actual)}:AA{str(pos_actual)}")
                    ws[f"Z{str(pos_actual)}"] = ws_datos[f"B{str(indice_datos+1)}"].value
                    ws[f"AB{str(pos_actual)}"] = valores[7]
                    ws.merge_cells(
                        f"AC{str(pos_actual)}:AH{str(pos_actual)}")
                    ws[f"AC{str(pos_actual)}"] = ws_datos[f"B{str(indice_datos+2)}"].value
                    ws[f"AC{str(pos_actual)}"].alignment = Alignment(
                        horizontal='center', vertical='center')
                # endif
                indice_datos += 4
            elif valores[5] == "TABLA":
                pos_actual += 1
                createMAXMIN(ws_source=ws_data,
                             ws_destination=ws_datos, idx=indice_datos)

                ws.merge_cells(f"J{str(pos_actual)}:O{str(pos_actual)}")
                ws[f"J{pos_actual}"] = valores[6] + " MAX"

                ws.merge_cells(f"P{str(pos_actual)}:U{str(pos_actual)}")
                ws[f"P{pos_actual}"] = "Fecha de maxima:"

                ws.merge_cells(f"V{str(pos_actual)}:AA{str(pos_actual)}")
                ws[f"V{pos_actual}"] = valores[6] + " MIN"

                ws.merge_cells(f"AB{str(pos_actual)}:AG{str(pos_actual)}")
                ws[f"AB{pos_actual}"] = "Fecha de minimo:"

                apply_style_to_range(
                    ws, f"J{str(pos_actual)}:AG{str(pos_actual)}")

                pos_actual += 1
                x = 0
                for col_idx, col in enumerate(ws_data.iter_cols(min_col=2)):
                    fill = PatternFill(start_color=series_colors_aux[(x) % len(
                        series_colors_aux)], end_color=series_colors_aux[(x) % len(series_colors_aux)], fill_type='solid')

                    ws.merge_cells(
                        f"A{str(pos_actual)}:I{str(pos_actual)}")
                    ws[f"A{str(pos_actual)}"] = str(col[0].value)
                    ws[f"A{str(pos_actual)}"].border = border
                    ws[f"A{str(pos_actual)}"].fill = fill

                    ws.merge_cells(
                        f"J{str(pos_actual)}:O{str(pos_actual)}")
                    ws[f"J{str(pos_actual)}"] = ws_datos[f"A{str(indice_datos+1)}"].value
                    ws[f"J{str(pos_actual)}"].border = border
                    ws[f"J{str(pos_actual)}"].fill = fill

                    ws.merge_cells(
                        f"P{str(pos_actual)}:U{str(pos_actual)}")
                    ws[f"P{str(pos_actual)}"] = ws_datos[f"A{str(indice_datos+2)}"].value
                    ws[f"P{str(pos_actual)}"].border = border
                    ws[f"P{str(pos_actual)}"].fill = fill

                    ws.merge_cells(
                        f"V{str(pos_actual)}:AA{str(pos_actual)}")
                    ws[f"V{str(pos_actual)}"] = ws_datos[f"B{str(indice_datos+1)}"].value
                    ws[f"V{str(pos_actual)}"].border = border
                    ws[f"V{str(pos_actual)}"].fill = fill

                    ws.merge_cells(
                        f"AB{str(pos_actual)}:AG{str(pos_actual)}")
                    ws[f"AB{str(pos_actual)}"] = ws_datos[f"B{str(indice_datos+2)}"].value
                    ws[f"AB{str(pos_actual)}"].border = border
                    ws[f"AB{str(pos_actual)}"].fill = fill

                    apply_style_to_range(
                        ws, f"A{str(pos_actual)}:AG{str(pos_actual)}")

                    x += 1
                    pos_actual += 1
                    indice_datos += 4
                # endfor
            caca = 10
            # endif
            series_colors = series_colors[1:] + series_colors[:1]
            series_colors_aux = series_colors_aux[1:] + series_colors_aux[:1]
        # endfor
    # endfor

    # Anadir pie de pagina
    ws.oddFooter.right.text = "Powered by \nALCORT INGENIERiA Y ASESORiA S.L."
    ws.oddFooter.center.text = "&[Page]"
    ws.evenFooter.right.text = "Powered by \nALCORT INGENIERiA Y ASESORiA S.L."
    ws.evenFooter.center.text = "&[Page]"

    # Guardar el archivo como informe_semanal_alamo_v2.xlsx
    wb.save(
        f"{INFORMES_DIR}{str(time_finish_aux.strftime('%Y-%m-%d'))}_informe_{TITULO}.xlsx")
    return str(f"{str(time_finish_aux.strftime('%Y-%m-%d'))}_informe_{TITULO}.xlsx")
# endfunction
##################################################################################################

def main():
    if ACTIVAR_SELECCION_RANGO_DE_FECHAS:
        mostrar_popup()
    # endif
    i = 0
    for clave, valor in DASHBOARDS.items():
        os.makedirs(DATA_DIR+"/" + clave[0]+"/", exist_ok=True)
        os.makedirs(INFORMES_DIR, exist_ok=True)
        obtenerDatosGrafana(
            DATA_DIR+"/"+clave[0]+"/", GRAFANA_URL, API_KEY, UIDS[i], valor)
        excelDeDatos(DATA_DIR+"/" + clave[0]+"/", valor)
        i += 1
    # endfor
    print(informe("Informe Semanal CT Cristo", DASHBOARDS))
# endfunction

if __name__ == "__main__":
    main()
# endif